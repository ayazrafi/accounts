import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
from PySide6.QtWidgets import QFileDialog, QMessageBox

def export_sales_register(parent_widget, company, vouchers, date_range_str):
    """
    Exports a Sales Register Excel file formatted like the sample provided.
    vouchers: list of voucher dicts, each should have 'invoice_items' and 'entries' (for taxes)
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Register"

        # --- Styles ---
        bold_font = Font(bold=True, size=11)
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # --- Header Section ---
        # 1. Company Name (ROYAL SONS (2025-2026))
        ws.merge_cells('A1:O1')
        ws['A1'] = f"{company.get('name', 'Company Name')} ({company.get('financial_year', '2025-2026')})"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_left

        # 2. Address line 1
        ws.merge_cells('A2:O2')
        ws['A2'] = company.get('address', '')
        ws['A2'].alignment = align_left

        # 3. Address line 2 / City
        ws.merge_cells('A3:O3')
        ws['A3'] = f"{company.get('city', '')} {company.get('state', '')}"
        ws['A3'].alignment = align_left
        
        # Border for header area bottom
        for col in range(1, 16):
            ws.cell(row=3, column=col).border = Border(bottom=Side(style='thin'))

        # 4. "Sales Register"
        ws.merge_cells('A4:O4')
        ws['A4'] = "Sales Register"
        ws['A4'].font = header_font
        ws['A4'].alignment = align_left

        # 5. Date Range
        ws.merge_cells('A5:O5')
        ws['A5'] = date_range_str
        ws['A5'].alignment = align_left

        # --- Table Headers ---
        # Row 6: Column titles
        headers = [
            "Date", "Particulars", "Voucher Type", "Voucher No.", "GSTIN/UIN", 
            "Sales Tax No.", "Quantity", "Value", "Gross Total", "Sale", 
            "CGST 9%", "SGST 9%", "ROUND OFF", "Cgst 14%", "Sgst 14%"
        ]
        
        # Actually, let's detect tax columns dynamically if possible, 
        # but for now we follow the exact requested format.
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = align_center
            cell.border = border_thin

        # --- Data Rows ---
        current_row = 7
        grand_total_value = 0
        grand_total_gross = 0
        grand_total_sale = 0
        grand_total_cgst9 = 0
        grand_total_sgst9 = 0
        grand_total_cgst14 = 0
        grand_total_sgst14 = 0
        grand_total_roundoff = 0

        for v in vouchers:
            # First row of voucher: Header info
            ws.cell(row=current_row, column=1, value=v.get('date', ''))
            ws.cell(row=current_row, column=2, value=v.get('party_name', 'Unknown Party')).font = bold_font
            ws.cell(row=current_row, column=3, value=v.get('voucher_type', 'Sales'))
            ws.cell(row=current_row, column=4, value=v.get('voucher_no', ''))
            ws.cell(row=current_row, column=5, value=v.get('gstin', ''))
            
            # Tax mapping
            cgst9 = sgst9 = cgst14 = sgst14 = roundoff = 0
            sale_value = 0 # Taxable value
            
            # Identify taxes from entries
            for entry in v.get('entries', []):
                lname = entry.get('ledger_name', '').upper()
                amt = abs(entry.get('amount', 0))
                if 'CGST' in lname and '9%' in lname: cgst9 += amt
                elif 'SGST' in lname and '9%' in lname: sgst9 += amt
                elif 'CGST' in lname and '14%' in lname: cgst14 += amt
                elif 'SGST' in lname and '14%' in lname: sgst14 += amt
                elif 'ROUND' in lname: roundoff += entry.get('amount', 0) # Could be negative
                # Sale value is usually the sum of items, or if we identify a 'Sales' ledger
                if 'SALES' in lname and 'CGST' not in lname and 'SGST' not in lname:
                    sale_value += amt

            gross_total = v.get('amount', 0)
            
            ws.cell(row=current_row, column=8, value=sale_value).font = bold_font
            ws.cell(row=current_row, column=9, value=gross_total).font = bold_font
            ws.cell(row=current_row, column=10, value=sale_value)
            ws.cell(row=current_row, column=11, value=cgst9)
            ws.cell(row=current_row, column=12, value=sgst9)
            ws.cell(row=current_row, column=13, value=roundoff)
            ws.cell(row=current_row, column=14, value=cgst14)
            ws.cell(row=current_row, column=15, value=sgst14)
            
            # Formatting for the header row
            for c in range(1, 16):
                ws.cell(row=current_row, column=c).border = border_thin
                if c >= 8: ws.cell(row=current_row, column=c).alignment = align_right

            voucher_start_row = current_row
            current_row += 1

            # Subsequent rows: Stock Items
            items = v.get('invoice_items', [])
            for item in items:
                # Format: "HSN Name" or just name
                p_text = f"   {item.get('hsn', '')} {item.get('item_name', '')}"
                ws.cell(row=current_row, column=2, value=p_text).font = Font(italic=True)
                
                qty_text = f"{item.get('qty', 0)} {item.get('unit', 'NOS')}"
                ws.cell(row=current_row, column=7, value=qty_text).alignment = align_right
                
                val = item.get('taxable_value', 0)
                ws.cell(row=current_row, column=8, value=val).alignment = align_right
                
                # Borders for item rows (sides only to match sample feel, or thin)
                for c in range(1, 16):
                    ws.cell(row=current_row, column=c).border = Border(left=Side(style='thin'), right=Side(style='thin'))
                
                current_row += 1

            # Accumulate totals
            grand_total_gross += gross_total
            grand_total_sale += sale_value
            grand_total_cgst9 += cgst9
            grand_total_sgst9 += sgst9
            grand_total_cgst14 += cgst14
            grand_total_sgst14 += sgst14
            grand_total_roundoff += roundoff

        # --- Grand Total Row ---
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "Grand Total"
        ws[f'A{current_row}'].font = bold_font
        ws[f'A{current_row}'].alignment = align_right
        
        totals = [
            (8, grand_total_sale), 
            (9, grand_total_gross), 
            (10, grand_total_sale), 
            (11, grand_total_cgst9), 
            (12, grand_total_sgst9), 
            (13, grand_total_roundoff),
            (14, grand_total_cgst14),
            (15, grand_total_sgst14)
        ]
        
        for col, val in totals:
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = bold_font
            cell.alignment = align_right
            cell.border = border_thin
        
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = border_thin

        # --- Final Polish ---
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 12
        for col in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws.column_dimensions[col].width = 14

        # --- Save File ---
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suggested_name = f"Sales_Register_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(parent_widget, "Save Sales Register", suggested_name, "Excel Files (*.xlsx)")
        
        if path:
            wb.save(path)
            QMessageBox.information(parent_widget, "Success", f"Sales Register exported successfully to:\n{path}")
            # Try to open the file
            try:
                os.startfile(path)
            except:
                pass

    except Exception as e:
        QMessageBox.critical(parent_widget, "Export Error", f"Failed to export Excel:\n{str(e)}")
        import traceback
        traceback.print_exc()
