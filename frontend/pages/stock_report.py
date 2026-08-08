import os
import calendar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
    QMessageBox, QDateEdit, QFrame, QFormLayout, QGridLayout,
    QScrollArea, QDialog, QComboBox
)
from PySide6.QtCore import Qt, QDate, QSize, QRectF
from PySide6.QtGui import QFont, QIcon, QPainter, QColor, QLinearGradient

import frontend.api_client as api
from frontend.utils import SearchableComboBox, DateEdit, get_icon, format_inr, format_indian_number
import frontend.session as session
from frontend.pages.voucher import InvoiceVoucherDialog, PaymentReceiptDialog, JournalVoucherDialog, INVOICE_TYPES


# ── Generic Excel Exporter ────────────────────────────────────────────────────
def export_table_to_excel(parent_widget, title_text, table_widget, date_range_str):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="1e3a5f")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="ffffff")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        data_font = Font(name="Segoe UI", size=11)
        bold_font = Font(name="Segoe UI", size=11, bold=True)
        border_thin = Border(left=Side(style='thin', color='e2e8f0'),
                             right=Side(style='thin', color='e2e8f0'),
                             top=Side(style='thin', color='e2e8f0'),
                             bottom=Side(style='thin', color='e2e8f0'))
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        
        co_name = session.company_name or "Company Name"
        
        ws.cell(row=1, column=1, value=co_name).font = title_font
        ws.cell(row=2, column=1, value=title_text).font = Font(name="Segoe UI", size=12, bold=True, color="475569")
        ws.cell(row=3, column=1, value=date_range_str).font = Font(name="Segoe UI", size=10, italic=True)
        
        headers = []
        for col in range(table_widget.columnCount()):
            headers.append(table_widget.horizontalHeaderItem(col).text())
            
        header_row = 5
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        current_row = 6
        for row in range(table_widget.rowCount()):
            first_item = table_widget.item(row, 0)
            is_total_row = (first_item is not None and "total" in first_item.text().lower())
                            
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                val = item.text() if item else ""
                
                clean_val = val.replace("₹", "").replace(",", "").replace(" ", "").strip()
                try:
                    if " " not in clean_val:
                        num_val = float(clean_val)
                        cell = ws.cell(row=current_row, column=col+1, value=num_val)
                    else:
                        cell = ws.cell(row=current_row, column=col+1, value=val)
                except ValueError:
                    cell = ws.cell(row=current_row, column=col+1, value=val)
                    
                cell.font = bold_font if is_total_row else data_font
                cell.border = border_thin
                
                if item:
                    align = item.textAlignment()
                    if align & Qt.AlignmentFlag.AlignRight:
                        cell.alignment = align_right
                    elif align & Qt.AlignmentFlag.AlignCenter:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                else:
                    cell.alignment = align_left
                    
            current_row += 1
            
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suggested_name = f"{title_text.replace(' ', '_')}_{timestamp}.xlsx"
        path, _ = QFileDialog.getSaveFileName(parent_widget, f"Save {title_text}", suggested_name, "Excel Files (*.xlsx)")
        
        if path:
            wb.save(path)
            QMessageBox.information(parent_widget, "Success", f"Exported successfully to:\n{path}")
            try:
                os.startfile(path)
            except:
                pass
    except Exception as e:
        QMessageBox.critical(parent_widget, "Export Error", f"Failed to export Excel:\n{str(e)}")


# ── Premium Bar Chart ─────────────────────────────────────────────────────────
class MonthlyBarChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(240)
        self._data = []  # list of {"month": str, "value": float}
        self._title = "Monthly Outwards"

    def setData(self, data, title="Monthly Outwards"):
        self._data = data
        self._title = title
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        w = self.width()
        h = self.height()
        
        painter.fillRect(0, 0, w, h, QColor("#ffffff"))
        
        left_m = 65
        right_m = 30
        top_m = 45
        bottom_m = 40
        
        plot_w = w - left_m - right_m
        plot_h = h - top_m - bottom_m
        
        # Title
        painter.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        painter.setPen(QColor("#1e293b"))
        painter.drawText(20, 25, self._title)
        
        if not self._data:
            painter.setFont(QFont("Segoe UI", 10))
            painter.setPen(QColor("#64748b"))
            painter.drawText(left_m, top_m + plot_h/2, "No monthly movement data to display")
            return
            
        max_val = max(d["value"] for d in self._data)
        if max_val <= 0:
            max_val = 1.0
            
        import math
        magnitude = 10**int(math.log10(max_val)) if max_val > 0 else 1
        if magnitude == 0: magnitude = 1
        max_val = math.ceil(max_val / magnitude) * magnitude
        
        # Draw horizontal gridlines
        painter.setFont(QFont("Segoe UI", 8))
        for i in range(5):
            y = top_m + plot_h - (i * plot_h / 4)
            val = (i * max_val / 4)
            painter.setPen(QColor("#f1f5f9"))
            painter.drawLine(left_m, y, left_m + plot_w, y)
            painter.setPen(QColor("#64748b"))
            val_str = f"{val:,.2f}" if val % 1 != 0 else f"{int(val):,}"
            painter.drawText(10, y + 4, val_str)
            
        # Draw Bars
        bar_count = len(self._data)
        col_w = plot_w / bar_count
        bar_w = col_w * 0.6
        
        for idx, d in enumerate(self._data):
            val = d["value"]
            m_name = d["month"]
            
            bx = left_m + (idx * col_w) + (col_w - bar_w) / 2
            bh = (val / max_val) * plot_h
            by = top_m + plot_h - bh
            
            # Gradient bar
            grad = QLinearGradient(bx, by, bx, by + bh)
            grad.setColorAt(0.0, QColor("#3b82f6"))
            grad.setColorAt(1.0, QColor("#93c5fd"))
            
            painter.setBrush(grad)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(QRectF(bx, by, bar_w, bh), 4, 4)
            
            # Value label
            if val > 0:
                painter.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
                painter.setPen(QColor("#0f172a"))
                val_str = f"{val:,.0f}" if val % 1 == 0 else f"{val:,.2f}"
                text_w = painter.fontMetrics().horizontalAdvance(val_str)
                painter.drawText(bx + (bar_w - text_w)/2, by - 6, val_str)
                
            # Month label
            painter.setFont(QFont("Segoe UI", 9))
            painter.setPen(QColor("#475569"))
            text_w = painter.fontMetrics().horizontalAdvance(m_name[:3])
            painter.drawText(bx + (bar_w - text_w)/2, top_m + plot_h + 20, m_name[:3])


# ── 1. Stock Group Summary Page ───────────────────────────────────────────────
class StockGroupSummaryPage(QWidget):
    def __init__(self):
        super().__init__()
        self._groups = []
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        hdr = QHBoxLayout()
        title = QLabel("Stock Group Summary")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a5f;")
        hdr.addWidget(title)
        hdr.addStretch()

        hdr.addWidget(QLabel("Stock Group:"))
        self.group_cb = SearchableComboBox()
        self.group_cb.setMinimumWidth(200)
        self.group_cb.currentIndexChanged.connect(self._load)
        hdr.addWidget(self.group_cb)

        hdr.addWidget(QLabel("As of:"))
        self.to_date = DateEdit(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.dateChanged.connect(self._load)
        hdr.addWidget(self.to_date)

        export_btn = QPushButton("  Export")
        export_btn.setIcon(get_icon("frontend/assets/icons/upload.svg", "#ffffff"))
        export_btn.setIconSize(QSize(16, 16))
        export_btn.clicked.connect(self._export)
        hdr.addWidget(export_btn)

        layout.addLayout(hdr)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Particulars", "Closing Quantity", "Closing Rate", "Closing Value"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_row_double_click)
        layout.addWidget(self.table)

    def showEvent(self, e):
        self._refresh_groups()
        super().showEvent(e)

    def _refresh_groups(self):
        self.group_cb.blockSignals(True)
        try:
            self._groups = api.list_stock_groups()
            self.group_cb.clear()
            self.group_cb.addItem("All Groups", None)
            for g in self._groups:
                self.group_cb.addItem(g["name"], g["_id"])
        except Exception:
            pass
        self.group_cb.blockSignals(False)
        self._load()

    def _load(self):
        self.table.setRowCount(0)
        gid = self.group_cb.currentData()
        to_dt = self.to_date.date().toString("yyyy-MM-dd")
        try:
            res = api.stock_group_summary(group_id=gid, to_date=to_dt)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return

        rows = res.get("rows", [])
        self.table.setSortingEnabled(False)
        for row, r in enumerate(rows):
            self.table.insertRow(row)
            
            p_item = QTableWidgetItem(r.get("name", ""))
            p_item.setData(Qt.ItemDataRole.UserRole, r.get("item_id"))
            self.table.setItem(row, 0, p_item)

            unit = r.get("unit", "")
            qty_text = f"{r.get('qty', 0):.3f} {unit}".strip()
            qty_item = QTableWidgetItem(qty_text)
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 1, qty_item)

            rate_item = QTableWidgetItem(format_indian_number(r.get("rate", 0.0)))
            rate_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 2, rate_item)

            val_item = QTableWidgetItem(format_indian_number(r.get("value", 0.0)))
            val_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 3, val_item)

        # Totals Row
        tot_row = len(rows)
        self.table.insertRow(tot_row)
        tot_p = QTableWidgetItem("Grand Total")
        tot_p.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(tot_row, 0, tot_p)

        tot_qty = QTableWidgetItem(f"{res.get('total_qty', 0):.3f}")
        tot_qty.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_qty.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 1, tot_qty)

        self.table.setItem(tot_row, 2, QTableWidgetItem("")) # empty rate for totals

        tot_val = QTableWidgetItem(format_indian_number(res.get("total_value", 0.0)))
        tot_val.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_val.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 3, tot_val)

        self.table.setSortingEnabled(True)

    def _on_row_double_click(self):
        row = self.table.currentRow()
        if row < 0 or row == self.table.rowCount() - 1:
            return
        item_cell = self.table.item(row, 0)
        if item_cell:
            item_id = item_cell.data(Qt.ItemDataRole.UserRole)
            if item_id:
                win = self.window()
                if win and hasattr(win, "_navigate_to"):
                    monthly_page = win.page_list[16]
                    monthly_page.set_item(item_id)
                    win._navigate_to(16)

    def _export(self):
        export_table_to_excel(
            self,
            "Stock Group Summary",
            self.table,
            f"As of {self.to_date.date().toString('yyyy-MM-dd')}"
        )


# ── 2. Stock Category Summary Page ────────────────────────────────────────────
class StockCategorySummaryPage(QWidget):
    def __init__(self):
        super().__init__()
        self._categories = []
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        hdr = QHBoxLayout()
        title = QLabel("Stock Category Summary")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a5f;")
        hdr.addWidget(title)
        hdr.addStretch()

        hdr.addWidget(QLabel("Stock Category:"))
        self.cat_cb = SearchableComboBox()
        self.cat_cb.setMinimumWidth(200)
        self.cat_cb.currentIndexChanged.connect(self._load)
        hdr.addWidget(self.cat_cb)

        hdr.addWidget(QLabel("As of:"))
        self.to_date = DateEdit(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.dateChanged.connect(self._load)
        hdr.addWidget(self.to_date)

        export_btn = QPushButton("  Export")
        export_btn.setIcon(get_icon("frontend/assets/icons/upload.svg", "#ffffff"))
        export_btn.setIconSize(QSize(16, 16))
        export_btn.clicked.connect(self._export)
        hdr.addWidget(export_btn)

        layout.addLayout(hdr)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Particulars", "Closing Quantity", "Closing Rate", "Closing Value"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_row_double_click)
        layout.addWidget(self.table)

    def showEvent(self, e):
        self._refresh_categories()
        super().showEvent(e)

    def _refresh_categories(self):
        self.cat_cb.blockSignals(True)
        try:
            self._categories = api.list_stock_categories()
            self.cat_cb.clear()
            self.cat_cb.addItem("All Categories", None)
            for c in self._categories:
                self.cat_cb.addItem(c["name"], c["_id"])
        except Exception:
            pass
        self.cat_cb.blockSignals(False)
        self._load()

    def _load(self):
        self.table.setRowCount(0)
        cid = self.cat_cb.currentData()
        to_dt = self.to_date.date().toString("yyyy-MM-dd")
        try:
            res = api.stock_category_summary(category_id=cid, to_date=to_dt)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return

        rows = res.get("rows", [])
        self.table.setSortingEnabled(False)
        for row, r in enumerate(rows):
            self.table.insertRow(row)
            
            p_item = QTableWidgetItem(r.get("name", ""))
            p_item.setData(Qt.ItemDataRole.UserRole, r.get("item_id"))
            self.table.setItem(row, 0, p_item)

            unit = r.get("unit", "")
            qty_text = f"{r.get('qty', 0):.3f} {unit}".strip()
            qty_item = QTableWidgetItem(qty_text)
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 1, qty_item)

            rate_item = QTableWidgetItem(format_indian_number(r.get("rate", 0.0)))
            rate_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 2, rate_item)

            val_item = QTableWidgetItem(format_indian_number(r.get("value", 0.0)))
            val_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 3, val_item)

        # Totals Row
        tot_row = len(rows)
        self.table.insertRow(tot_row)
        tot_p = QTableWidgetItem("Grand Total")
        tot_p.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(tot_row, 0, tot_p)

        tot_qty = QTableWidgetItem(f"{res.get('total_qty', 0):.3f}")
        tot_qty.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_qty.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 1, tot_qty)

        self.table.setItem(tot_row, 2, QTableWidgetItem(""))

        tot_val = QTableWidgetItem(format_indian_number(res.get("total_value", 0.0)))
        tot_val.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_val.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 3, tot_val)

        self.table.setSortingEnabled(True)

    def _on_row_double_click(self):
        row = self.table.currentRow()
        if row < 0 or row == self.table.rowCount() - 1:
            return
        item_cell = self.table.item(row, 0)
        if item_cell:
            item_id = item_cell.data(Qt.ItemDataRole.UserRole)
            if item_id:
                win = self.window()
                if win and hasattr(win, "_navigate_to"):
                    monthly_page = win.page_list[16]
                    monthly_page.set_item(item_id)
                    win._navigate_to(16)

    def _export(self):
        export_table_to_excel(
            self,
            "Stock Category Summary",
            self.table,
            f"As of {self.to_date.date().toString('yyyy-MM-dd')}"
        )


# ── 3. Stock Item Monthly Summary Page ────────────────────────────────────────
class StockMonthlySummaryPage(QWidget):
    def __init__(self):
        super().__init__()
        self._items = []
        self._current_item_id = None
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(12)

        hdr = QHBoxLayout()
        title = QLabel("Stock Item Monthly Summary")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a5f;")
        hdr.addWidget(title)
        hdr.addStretch()

        hdr.addWidget(QLabel("Stock Item:"))
        self.item_cb = SearchableComboBox()
        self.item_cb.setMinimumWidth(240)
        self.item_cb.currentIndexChanged.connect(self._on_item_changed)
        hdr.addWidget(self.item_cb)

        hdr.addWidget(QLabel("From:"))
        self.from_date = DateEdit(QDate(QDate.currentDate().year(), 4, 1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("yyyy-MM-dd")
        self.from_date.dateChanged.connect(self._load)
        hdr.addWidget(self.from_date)

        hdr.addWidget(QLabel("To:"))
        self.to_date = DateEdit(QDate(QDate.currentDate().year() + 1, 3, 31))
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.dateChanged.connect(self._load)
        hdr.addWidget(self.to_date)

        export_btn = QPushButton("  Export")
        export_btn.setIcon(get_icon("frontend/assets/icons/upload.svg", "#ffffff"))
        export_btn.setIconSize(QSize(16, 16))
        export_btn.clicked.connect(self._export)
        hdr.addWidget(export_btn)

        layout.addLayout(hdr)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels([
            "Particulars", "Inwards Qty", "Inwards Value",
            "Outwards Qty", "Outwards Value", "Closing Qty", "Closing Value"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_row_double_click)
        layout.addWidget(self.table)

        # Bar chart section
        self.chart = MonthlyBarChart()
        layout.addWidget(self.chart, 1)

    def set_item(self, item_id):
        self._current_item_id = item_id
        # Will trigger refresh_items and load
        
    def showEvent(self, e):
        self._refresh_items()
        super().showEvent(e)

    def _refresh_items(self):
        self.item_cb.blockSignals(True)
        try:
            self._items = api.list_stock_items()
            self.item_cb.clear()
            for i in self._items:
                self.item_cb.addItem(i["name"], i["_id"])
            if self._current_item_id:
                self.item_cb.setCurrentData(self._current_item_id)
        except Exception:
            pass
        self.item_cb.blockSignals(False)
        self._load()

    def _on_item_changed(self):
        self._current_item_id = self.item_cb.currentData()
        self._load()

    def _load(self):
        self.table.setRowCount(0)
        self.chart.setData([])
        iid = self._current_item_id or self.item_cb.currentData()
        if not iid:
            return
            
        f_dt = self.from_date.date().toString("yyyy-MM-dd")
        t_dt = self.to_date.date().toString("yyyy-MM-dd")
        
        try:
            res = api.stock_monthly_summary(item_id=iid, from_date=f_dt, to_date=t_dt)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return

        unit = res.get("unit", "")
        op = res.get("opening_balance", {})
        monthly = res.get("monthly", [])
        
        # 1. Opening Balance Row
        self.table.insertRow(0)
        op_part = QTableWidgetItem("Opening Balance")
        op_part.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(0, 0, op_part)
        for c in range(1, 5):
            self.table.setItem(0, c, QTableWidgetItem(""))
            
        op_qty_txt = f"{op.get('qty', 0):.3f} {unit}".strip()
        op_qty_item = QTableWidgetItem(op_qty_txt)
        op_qty_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        op_qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(0, 5, op_qty_item)

        op_val_item = QTableWidgetItem(format_indian_number(op.get("val", 0.0)))
        op_val_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        op_val_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(0, 6, op_val_item)

        # 2. Monthly Rows
        total_in_qty = 0.0
        total_in_val = 0.0
        total_out_qty = 0.0
        total_out_val = 0.0
        chart_data = []

        for idx, m in enumerate(monthly, 1):
            self.table.insertRow(idx)
            
            # Particulars cell
            part_item = QTableWidgetItem(m.get("particulars", ""))
            part_item.setData(Qt.ItemDataRole.UserRole + 1, m.get("month_num"))
            part_item.setData(Qt.ItemDataRole.UserRole + 2, m.get("year"))
            self.table.setItem(idx, 0, part_item)

            # Inwards
            in_q = m.get("inwards_qty", 0.0)
            in_v = m.get("inwards_val", 0.0)
            total_in_qty += in_q
            total_in_val += in_v
            
            in_q_item = QTableWidgetItem(f"{in_q:.3f} {unit}".strip() if in_q > 0 else "")
            in_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 1, in_q_item)

            in_v_item = QTableWidgetItem(format_indian_number(in_v) if in_v > 0 else "")
            in_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 2, in_v_item)

            # Outwards
            out_q = m.get("outwards_qty", 0.0)
            out_v = m.get("outwards_val", 0.0)
            total_out_qty += out_q
            total_out_val += out_v
            
            out_q_item = QTableWidgetItem(f"{out_q:.3f} {unit}".strip() if out_q > 0 else "")
            out_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 3, out_q_item)

            out_v_item = QTableWidgetItem(format_indian_number(out_v) if out_v > 0 else "")
            out_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 4, out_v_item)

            # Closing
            cl_q = m.get("closing_qty", 0.0)
            cl_v = m.get("closing_val", 0.0)
            
            cl_q_item = QTableWidgetItem(f"{cl_q:.3f} {unit}".strip())
            cl_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 5, cl_q_item)

            cl_v_item = QTableWidgetItem(format_indian_number(cl_v))
            cl_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 6, cl_v_item)

            # Populate chart values (using Outwards quantity to match design)
            chart_data.append({"month": m.get("particulars", "")[:3], "value": out_q})

        # 3. Grand Total Row
        tot_row = len(monthly) + 1
        self.table.insertRow(tot_row)
        tot_part = QTableWidgetItem("Grand Total")
        tot_part.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(tot_row, 0, tot_part)

        tot_in_q_item = QTableWidgetItem(f"{total_in_qty:.3f} {unit}".strip() if total_in_qty > 0 else "")
        tot_in_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_in_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 1, tot_in_q_item)

        tot_in_v_item = QTableWidgetItem(format_indian_number(total_in_val) if total_in_val > 0 else "")
        tot_in_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_in_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 2, tot_in_v_item)

        tot_out_q_item = QTableWidgetItem(f"{total_out_qty:.3f} {unit}".strip() if total_out_qty > 0 else "")
        tot_out_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_out_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 3, tot_out_q_item)

        tot_out_v_item = QTableWidgetItem(format_indian_number(total_out_val) if total_out_val > 0 else "")
        tot_out_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_out_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 4, tot_out_v_item)

        # Closing total matches final month's closing
        final_cl_q = monthly[-1].get("closing_qty", 0.0) if monthly else op.get("qty", 0.0)
        final_cl_v = monthly[-1].get("closing_val", 0.0) if monthly else op.get("val", 0.0)

        tot_cl_q_item = QTableWidgetItem(f"{final_cl_q:.3f} {unit}".strip())
        tot_cl_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_cl_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 5, tot_cl_q_item)

        tot_cl_v_item = QTableWidgetItem(format_indian_number(final_cl_v))
        tot_cl_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_cl_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 6, tot_cl_v_item)

        # Refresh custom bar chart
        self.chart.setData(chart_data, f"Outwards Quantity Summary for {res.get('item_name', '')}")

    def _on_row_double_click(self):
        row = self.table.currentRow()
        if row <= 0 or row == self.table.rowCount() - 1:
            return  # skip opening and totals
            
        part_cell = self.table.item(row, 0)
        if part_cell:
            m_num = part_cell.data(Qt.ItemDataRole.UserRole + 1)
            year = part_cell.data(Qt.ItemDataRole.UserRole + 2)
            if m_num and year and self._current_item_id:
                last_day = calendar.monthrange(year, m_num)[1]
                start_dt = f"{year}-{m_num:02d}-01"
                end_dt = f"{year}-{m_num:02d}-{last_day:02d}"
                
                win = self.window()
                if win and hasattr(win, "_navigate_to"):
                    vch_page = win.page_list[17]
                    vch_page.set_filter(self._current_item_id, start_dt, end_dt)
                    win._navigate_to(17)

    def _export(self):
        export_table_to_excel(
            self,
            "Stock Item Monthly Summary",
            self.table,
            f"Period {self.from_date.date().toString('yyyy-MM-dd')} to {self.to_date.date().toString('yyyy-MM-dd')}"
        )


# ── 4. Stock Item Vouchers Page ───────────────────────────────────────────────
class StockVouchersPage(QWidget):
    def __init__(self):
        super().__init__()
        self._items = []
        self._current_item_id = None
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(16)

        hdr = QHBoxLayout()
        title = QLabel("Stock Item Vouchers")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a5f;")
        hdr.addWidget(title)
        hdr.addStretch()

        hdr.addWidget(QLabel("Stock Item:"))
        self.item_cb = SearchableComboBox()
        self.item_cb.setMinimumWidth(240)
        self.item_cb.currentIndexChanged.connect(self._on_item_changed)
        hdr.addWidget(self.item_cb)

        hdr.addWidget(QLabel("From:"))
        self.from_date = DateEdit(QDate(QDate.currentDate().year(), 4, 1))
        self.from_date.setCalendarPopup(True)
        self.from_date.setDisplayFormat("yyyy-MM-dd")
        self.from_date.dateChanged.connect(self._load)
        hdr.addWidget(self.from_date)

        hdr.addWidget(QLabel("To:"))
        self.to_date = DateEdit(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        self.to_date.setDisplayFormat("yyyy-MM-dd")
        self.to_date.dateChanged.connect(self._load)
        hdr.addWidget(self.to_date)

        export_btn = QPushButton("  Export")
        export_btn.setIcon(get_icon("frontend/assets/icons/upload.svg", "#ffffff"))
        export_btn.setIconSize(QSize(16, 16))
        export_btn.clicked.connect(self._export)
        hdr.addWidget(export_btn)

        layout.addLayout(hdr)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels([
            "Date", "Particulars", "Vch Type", "Vch No",
            "Inwards Qty", "Inwards Value", "Outwards Qty", "Outwards Value",
            "Closing Qty", "Closing Value"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_row_double_click)
        layout.addWidget(self.table)

    def set_filter(self, item_id, start_date, end_date):
        self._current_item_id = item_id
        self.from_date.setDate(QDate.fromString(start_date, "yyyy-MM-dd"))
        self.to_date.setDate(QDate.fromString(end_date, "yyyy-MM-dd"))
        # load will trigger via date changes and item dropdown selections

    def showEvent(self, e):
        self._refresh_items()
        super().showEvent(e)

    def _refresh_items(self):
        self.item_cb.blockSignals(True)
        try:
            self._items = api.list_stock_items()
            self.item_cb.clear()
            for i in self._items:
                self.item_cb.addItem(i["name"], i["_id"])
            if self._current_item_id:
                self.item_cb.setCurrentData(self._current_item_id)
        except Exception:
            pass
        self.item_cb.blockSignals(False)
        self._load()

    def _on_item_changed(self):
        self._current_item_id = self.item_cb.currentData()
        self._load()

    def _load(self):
        self.table.setRowCount(0)
        iid = self._current_item_id or self.item_cb.currentData()
        if not iid:
            return
            
        f_dt = self.from_date.date().toString("yyyy-MM-dd")
        t_dt = self.to_date.date().toString("yyyy-MM-dd")
        
        try:
            res = api.stock_item_vouchers(item_id=iid, from_date=f_dt, to_date=t_dt)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return

        unit = res.get("unit", "")
        op = res.get("opening_balance", {})
        rows = res.get("rows", [])
        
        self.table.setSortingEnabled(False)
        
        # 1. Opening Balance Row
        self.table.insertRow(0)
        self.table.setItem(0, 0, QTableWidgetItem(""))
        
        op_part = QTableWidgetItem("Opening Balance")
        op_part.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(0, 1, op_part)
        
        for c in range(2, 8):
            self.table.setItem(0, c, QTableWidgetItem(""))
            
        op_qty_txt = f"{op.get('qty', 0):.3f} {unit}".strip()
        op_qty_item = QTableWidgetItem(op_qty_txt)
        op_qty_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        op_qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(0, 8, op_qty_item)

        op_val_item = QTableWidgetItem(format_indian_number(op.get("val", 0.0)))
        op_val_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        op_val_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(0, 9, op_val_item)

        # 2. Transaction Rows
        total_in_qty = 0.0
        total_in_val = 0.0
        total_out_qty = 0.0
        total_out_val = 0.0

        for idx, r in enumerate(rows, 1):
            self.table.insertRow(idx)
            
            # Date cell stores voucher details
            dt_item = QTableWidgetItem(r.get("date", ""))
            dt_item.setData(Qt.ItemDataRole.UserRole, r.get("voucher_id"))
            dt_item.setData(Qt.ItemDataRole.UserRole + 1, r.get("voucher_type"))
            self.table.setItem(idx, 0, dt_item)

            self.table.setItem(idx, 1, QTableWidgetItem(r.get("particulars", "")))
            self.table.setItem(idx, 2, QTableWidgetItem(r.get("voucher_type", "")))
            self.table.setItem(idx, 3, QTableWidgetItem(r.get("voucher_no", "")))

            # Inwards
            in_q = r.get("inwards_qty", 0.0)
            in_v = r.get("inwards_val", 0.0)
            total_in_qty += in_q
            total_in_val += in_v
            
            in_q_item = QTableWidgetItem(f"{in_q:.3f} {unit}".strip() if in_q > 0 else "")
            in_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 4, in_q_item)

            in_v_item = QTableWidgetItem(format_indian_number(in_v) if in_v > 0 else "")
            in_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 5, in_v_item)

            # Outwards
            out_q = r.get("outwards_qty", 0.0)
            out_v = r.get("outwards_val", 0.0)
            total_out_qty += out_q
            total_out_val += out_v
            
            out_q_item = QTableWidgetItem(f"{out_q:.3f} {unit}".strip() if out_q > 0 else "")
            out_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 6, out_q_item)

            out_v_item = QTableWidgetItem(format_indian_number(out_v) if out_v > 0 else "")
            out_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 7, out_v_item)

            # Closing
            cl_q = r.get("closing_qty", 0.0)
            cl_v = r.get("closing_val", 0.0)
            
            cl_q_item = QTableWidgetItem(f"{cl_q:.3f} {unit}".strip())
            cl_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 8, cl_q_item)

            cl_v_item = QTableWidgetItem(format_indian_number(cl_v))
            cl_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(idx, 9, cl_v_item)

        # 3. Grand Total Row
        tot_row = len(rows) + 1
        self.table.insertRow(tot_row)
        self.table.setItem(tot_row, 0, QTableWidgetItem(""))
        
        tot_part = QTableWidgetItem("Grand Total")
        tot_part.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.table.setItem(tot_row, 1, tot_part)

        for c in range(2, 4):
            self.table.setItem(tot_row, c, QTableWidgetItem(""))

        tot_in_q_item = QTableWidgetItem(f"{total_in_qty:.3f} {unit}".strip() if total_in_qty > 0 else "")
        tot_in_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_in_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 4, tot_in_q_item)

        tot_in_v_item = QTableWidgetItem(format_indian_number(total_in_val) if total_in_val > 0 else "")
        tot_in_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_in_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 5, tot_in_v_item)

        tot_out_q_item = QTableWidgetItem(f"{total_out_qty:.3f} {unit}".strip() if total_out_qty > 0 else "")
        tot_out_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_out_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 6, tot_out_q_item)

        tot_out_v_item = QTableWidgetItem(format_indian_number(total_out_val) if total_out_val > 0 else "")
        tot_out_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_out_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 7, tot_out_v_item)

        final_cl_q = rows[-1].get("closing_qty", 0.0) if rows else op.get("qty", 0.0)
        final_cl_v = rows[-1].get("closing_val", 0.0) if rows else op.get("val", 0.0)

        tot_cl_q_item = QTableWidgetItem(f"{final_cl_q:.3f} {unit}".strip())
        tot_cl_q_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_cl_q_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 8, tot_cl_q_item)

        tot_cl_v_item = QTableWidgetItem(format_indian_number(final_cl_v))
        tot_cl_v_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        tot_cl_v_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(tot_row, 9, tot_cl_v_item)

        self.table.setSortingEnabled(True)

    def _on_row_double_click(self):
        row = self.table.currentRow()
        if row <= 0 or row == self.table.rowCount() - 1:
            return  # skip opening and totals
            
        dt_cell = self.table.item(row, 0)
        if dt_cell:
            vid = dt_cell.data(Qt.ItemDataRole.UserRole)
            vtype = dt_cell.data(Qt.ItemDataRole.UserRole + 1)
            if vid and vtype:
                self._edit_voucher(vid, vtype)

    def _edit_voucher(self, vid, vtype):
        if not session.has_permission(vtype, "update"):
            QMessageBox.warning(self, "Permission Denied", f"You do not have permission to edit {vtype} vouchers.")
            return
            
        try:
            voucher = api.get_voucher(vid)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return
            
        if vtype == "Credit Note":
            try: voucher["invoice_items"] = api.get_voucher_stock_txns(vid)
            except: pass
            self.window().open_credit_note(voucher)
            self._load()
            return
        elif vtype == "Debit Note":
            try: voucher["invoice_items"] = api.get_voucher_stock_txns(vid)
            except: pass
            self.window().open_debit_note(voucher)
            self._load()
            return
        elif vtype in INVOICE_TYPES:
            try: voucher["invoice_items"] = api.get_voucher_stock_txns(vid)
            except: pass
            dlg = InvoiceVoucherDialog(self, vtype, existing=voucher)
        elif vtype in ["Payment", "Receipt"]:
            dlg = PaymentReceiptDialog(self, vtype, existing=voucher)
        else:
            dlg = JournalVoucherDialog(self, vtype, existing=voucher)
            
        if dlg.exec():
            data = dlg.get_data()
            entries = data.get("entries", [])
            if not entries:
                QMessageBox.warning(self, "Error", "No entries.")
                return
            try:
                update_payload = {
                    "date": data["date"],
                    "narration": data["narration"],
                    "entries": entries,
                    "invoice_items": data.get("invoice_items")
                }
                if "grand_total" in data: update_payload["grand_total"] = data["grand_total"]
                if "linking" in data: update_payload["linking"] = data["linking"]
                api.update_voucher(vid, update_payload)
                self._load()
            except Exception as ex:
                QMessageBox.warning(self, "Error", str(ex))

    def _export(self):
        export_table_to_excel(
            self,
            "Stock Item Vouchers",
            self.table,
            f"Period {self.from_date.date().toString('yyyy-MM-dd')} to {self.to_date.date().toString('yyyy-MM-dd')}"
        )


# ── 5. Stock Query Page ───────────────────────────────────────────────────────
class StockQueryPage(QWidget):
    def __init__(self):
        super().__init__()
        self._items = []
        self._current_item_id = None
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(12)

        # Header Selector
        hdr = QHBoxLayout()
        title = QLabel("Stock Query")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1e3a5f;")
        hdr.addWidget(title)
        hdr.addStretch()

        hdr.addWidget(QLabel("Select Stock Item:"))
        self.item_cb = SearchableComboBox()
        self.item_cb.setMinimumWidth(300)
        self.item_cb.currentIndexChanged.connect(self._on_item_changed)
        hdr.addWidget(self.item_cb)
        layout.addLayout(hdr)

        # Scrollable dashboard content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content = QWidget()
        self.grid = QGridLayout(content)
        self.grid.setContentsMargins(0, 0, 0, 0)
        self.grid.setSpacing(12)

        # Header Card: Metadata
        self.meta_frame = QFrame()
        self.meta_frame.setStyleSheet("QFrame { background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 8px; }")
        self.meta_lay = QGridLayout(self.meta_frame)
        self.meta_lay.setContentsMargins(16, 16, 16, 16)
        self.meta_lay.setSpacing(8)
        self.grid.addWidget(self.meta_frame, 0, 0, 1, 2)

        # Left Table: Purchases
        self.purch_frame = QFrame()
        self.purch_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; }")
        purch_lay = QVBoxLayout(self.purch_frame)
        purch_lay.setContentsMargins(12, 12, 12, 12)
        
        self.purch_hdr = QLabel("<b>Purchases</b>")
        self.purch_hdr.setStyleSheet("color: #1e3a5f; font-size: 13px;")
        purch_lay.addWidget(self.purch_hdr)

        self.purch_table = QTableWidget(0, 5)
        self.purch_table.setHorizontalHeaderLabels(["Date", "Party Name", "Quantity", "Rate", "Amount"])
        self.purch_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.purch_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.purch_table.verticalHeader().setVisible(False)
        self.purch_table.setMinimumHeight(180)
        purch_lay.addWidget(self.purch_table)
        self.grid.addWidget(self.purch_frame, 1, 0)

        # Right Table: Sales
        self.sales_frame = QFrame()
        self.sales_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; }")
        sales_lay = QVBoxLayout(self.sales_frame)
        sales_lay.setContentsMargins(12, 12, 12, 12)

        self.sales_hdr = QLabel("<b>Sales</b>")
        self.sales_hdr.setStyleSheet("color: #065f46; font-size: 13px;")
        sales_lay.addWidget(self.sales_hdr)

        self.sales_table = QTableWidget(0, 5)
        self.sales_table.setHorizontalHeaderLabels(["Date", "Party Name", "Quantity", "Rate", "Amount"])
        self.sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.sales_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.sales_table.verticalHeader().setVisible(False)
        self.sales_table.setMinimumHeight(180)
        sales_lay.addWidget(self.sales_table)
        self.grid.addWidget(self.sales_frame, 1, 1)

        # Bottom Left: Godown details
        self.godown_frame = QFrame()
        self.godown_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; }")
        godown_lay = QVBoxLayout(self.godown_frame)
        godown_lay.setContentsMargins(12, 12, 12, 12)
        godown_lay.addWidget(QLabel("<b>Godown / Batch Details</b>"))
        
        self.godown_table = QTableWidget(0, 3)
        self.godown_table.setHorizontalHeaderLabels(["Godown", "Batch", "Quantity"])
        self.godown_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.godown_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.godown_table.verticalHeader().setVisible(False)
        self.godown_table.setMinimumHeight(150)
        godown_lay.addWidget(self.godown_table)
        self.grid.addWidget(self.godown_frame, 2, 0)

        # Bottom Right: Sibling items
        self.sibling_frame = QFrame()
        self.sibling_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #cbd5e1; border-radius: 8px; }")
        sib_lay = QVBoxLayout(self.sibling_frame)
        sib_lay.setContentsMargins(12, 12, 12, 12)
        sib_lay.addWidget(QLabel("<b>Items of Same Category</b>"))

        self.sibling_table = QTableWidget(0, 4)
        self.sibling_table.setHorizontalHeaderLabels(["Item Name", "Quantity", "Cost Price", "Sale Price"])
        self.sibling_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.sibling_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.sibling_table.verticalHeader().setVisible(False)
        self.sibling_table.setMinimumHeight(150)
        sib_lay.addWidget(self.sibling_table)
        self.grid.addWidget(self.sibling_frame, 2, 1)

        scroll.setWidget(content)
        layout.addWidget(scroll, 1)

    def showEvent(self, e):
        self._refresh_items()
        super().showEvent(e)

    def _refresh_items(self):
        self.item_cb.blockSignals(True)
        try:
            self._items = api.list_stock_items()
            self.item_cb.clear()
            for i in self._items:
                self.item_cb.addItem(i["name"], i["_id"])
            if self._current_item_id:
                self.item_cb.setCurrentData(self._current_item_id)
        except Exception:
            pass
        self.item_cb.blockSignals(False)
        self._load()

    def _on_item_changed(self):
        self._current_item_id = self.item_cb.currentData()
        self._load()

    def _load(self):
        # Clear UI
        for i in reversed(range(self.meta_lay.count())): 
            widget = self.meta_lay.itemAt(i).widget()
            if widget: widget.deleteLater()
            
        self.purch_table.setRowCount(0)
        self.sales_table.setRowCount(0)
        self.godown_table.setRowCount(0)
        self.sibling_table.setRowCount(0)

        iid = self._current_item_id or self.item_cb.currentData()
        if not iid:
            return
            
        try:
            res = api.stock_query(item_id=iid)
        except Exception as ex:
            QMessageBox.warning(self, "Error", str(ex))
            return

        details = res.get("item_details", {})
        closing = res.get("closing_balance", {})
        unit = details.get("unit", "Nos")

        # 1. Populate Metadata header grid
        metadata_fields = [
            ("Name", details.get("name", "")),
            ("Part No.", details.get("part_no", "-")),
            ("Group", details.get("group", "")),
            ("Category", details.get("category", "")),
            ("Closing Balance", f"{closing.get('qty', 0.0):.3f} {unit}"),
            ("Closing Value", format_inr(closing.get("val", 0.0))),
            ("Cost price", f"{format_inr(details.get('standard_cost', 0.0))}/{unit}"),
            ("Standard selling price", f"{format_inr(details.get('standard_selling_price', 0.0))}/{unit}"),
            ("Costing method", details.get("costing_method", "Avg. Cost")),
            ("Market valuation method", details.get("market_valuation_method", "Avg. Price")),
        ]

        # Draw details in 2 columns
        for idx, (label, val) in enumerate(metadata_fields):
            r = idx // 2
            c = (idx % 2) * 2
            
            lbl_widget = QLabel(f"<b>{label}:</b>")
            lbl_widget.setStyleSheet("color: #64748b; font-size: 12px;")
            val_widget = QLabel(str(val))
            val_widget.setStyleSheet("color: #0f172a; font-size: 13px; font-weight: bold;")
            
            self.meta_lay.addWidget(lbl_widget, r, c)
            self.meta_lay.addWidget(val_widget, r, c + 1)

        # 2. Last purchased / Last sold subtitle headers
        lp = res.get("last_purchased")
        if lp:
            self.purch_hdr.setText(f"<b>Purchases</b> (Last purchased on: {lp.get('date')}  {lp.get('party_name')}  {lp.get('qty')} {unit} @ {format_inr(lp.get('rate'))})")
        else:
            self.purch_hdr.setText("<b>Purchases</b> (No purchase history)")

        ls = res.get("last_sold")
        if ls:
            self.sales_hdr.setText(f"<b>Sales</b> (Last sold on: {ls.get('date')}  {ls.get('party_name')}  {ls.get('qty')} {unit} @ {format_inr(ls.get('rate'))})")
        else:
            self.sales_hdr.setText("<b>Sales</b> (No sales history)")

        # 3. Populate Purchases
        for row, p in enumerate(res.get("purchases", [])):
            self.purch_table.insertRow(row)
            self.purch_table.setItem(row, 0, QTableWidgetItem(p.get("date", "")))
            self.purch_table.setItem(row, 1, QTableWidgetItem(p.get("party_name", "")))
            
            qty_item = QTableWidgetItem(f"{p.get('qty', 0.0):.3f} {unit}".strip())
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.purch_table.setItem(row, 2, qty_item)

            rate_item = QTableWidgetItem(format_indian_number(p.get("rate", 0.0)))
            rate_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.purch_table.setItem(row, 3, rate_item)

            amt_item = QTableWidgetItem(format_indian_number(p.get("amount", 0.0)))
            amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.purch_table.setItem(row, 4, amt_item)

        # 4. Populate Sales
        for row, s in enumerate(res.get("sales", [])):
            self.sales_table.insertRow(row)
            self.sales_table.setItem(row, 0, QTableWidgetItem(s.get("date", "")))
            self.sales_table.setItem(row, 1, QTableWidgetItem(s.get("party_name", "")))

            qty_item = QTableWidgetItem(f"{s.get('qty', 0.0):.3f} {unit}".strip())
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sales_table.setItem(row, 2, qty_item)

            rate_item = QTableWidgetItem(format_indian_number(s.get("rate", 0.0)))
            rate_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sales_table.setItem(row, 3, rate_item)

            amt_item = QTableWidgetItem(format_indian_number(s.get("amount", 0.0)))
            amt_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sales_table.setItem(row, 4, amt_item)

        # 5. Populate Godowns
        self.godown_table.insertRow(0)
        self.godown_table.setItem(0, 0, QTableWidgetItem("Main Location"))
        self.godown_table.setItem(0, 1, QTableWidgetItem("Primary Batch"))
        
        qty_item = QTableWidgetItem(f"{closing.get('qty', 0.0):.3f} {unit}".strip())
        qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.godown_table.setItem(0, 2, qty_item)

        # 6. Populate sibling category items
        for row, sib in enumerate(res.get("same_category_items", [])):
            self.sibling_table.insertRow(row)
            self.sibling_table.setItem(row, 0, QTableWidgetItem(sib.get("name", "")))
            
            sib_unit = sib.get("unit", "Nos")
            qty_item = QTableWidgetItem(f"{sib.get('qty', 0.0):.3f} {sib_unit}".strip())
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sibling_table.setItem(row, 1, qty_item)

            cost_item = QTableWidgetItem(format_indian_number(sib.get("cost", 0.0)))
            cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sibling_table.setItem(row, 2, cost_item)

            sale_item = QTableWidgetItem(format_indian_number(sib.get("sale_price", 0.0)))
            sale_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.sibling_table.setItem(row, 3, sale_item)
